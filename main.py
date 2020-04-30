import os
import os.path
import speech_recognition as sr 
import time 
import random
import playsound
from gtts import gTTS
import openpyxl
from openpyxl import load_workbook
#from string import ascii_uppercase

os.system('cls')


rec = sr.Recognizer()
workbookTitel = "FitnessApp-Daten.xlsx"
sheetStatistiken = "Trainingsstatistiken"
sheetÜbungen = "Uebungen"

def erstnutzungPrüfen():


    if os.path.isfile(workbookTitel):
        athenaSagt('Willkommen zurück. Wollen wir wieder trainieren?')
    else:
        neueExcelTabelleErstellen()
        athenaSagt('Willkommen zur fitness app. Mein Name ist Athena, und ich werde dein Training begleiten.')
        # athenaSagt('Da du die App anscheinend zum ersten mal startest muss ich wissen welche Übungen und wie viele Sätze du mit wie vielen Wiederholungen machen willst')
        # athenaSagt('Schaue dazu auf dein Gerät und füge Übungen hinzu.')
def nutzerSagt():

    with sr.Microphone() as source:

        vInput = rec.listen(source)

        try: 
            vOutput = rec.recognize_google(vInput, language='de_DE')
        except sr.UnknownValueError:
            athenaSagt('Sorry, das habe ich nicht verstanden')
        except sr.RequestError:
            athenaSagt('Sorry, ich kann momentan auf den sprach service nicht zugreifen.')
        return vOutput
def athenaEntscheidet(vInput):
    if 'starte training' in vInput:
        athenaSagt('Ok, starte die Fitness App')
        starteTraining()    
    elif 'jo' in vInput:
        athenaSagt('Ok, starte die Fitness App')
    elif 'beenden' in vInput:
        athenaSagt('Alles klar. Bis zum nächsten mal.')
        exit()
    elif 'Übung hinzufügen' in vInput:
        übungHinzufügen()

def athenaSagt(_string):
    tts = gTTS(text=_string, lang='de')
    r = random.randint(1,10000000)
    audioFile = 'audio-' + str(r) + '.mp3'
    tts.save(audioFile)
    print(_string)
    playsound.playsound(audioFile)
    os.remove(audioFile)
def neueExcelTabelleErstellen():
    #Leeren Workbook erstellen
    wb = openpyxl.Workbook()

    #Sheet für Übungen und Fortschritt erstellen
    übungsSheet = wb.active
    übungsSheet.title = sheetÜbungen
    datenSheet = wb.create_sheet(sheetStatistiken)

    #Beim erstmaligen Erstellen des Excels noch nur ges. Wiederholungen. Beim Verarbeiten durch Athena müssen dann noch die einzelnen Wiederholungen der Sätze dazukommen.
    headerÜbungsSheet = ['Uebung', 'Saetze', 'Wiederholungen']
    headerDatenSheet = ['Datum']    

    # Überschriften in die oberste Zeile der Exceltabelle eintragen
    for i in range(1, len(headerÜbungsSheet)+1):
         übungsSheet.cell(row =1, column=i).value = headerÜbungsSheet[i-1]
    
    for i in range(1, len(headerDatenSheet)+1):
        datenSheet.cell(row =1, column=i).value = headerDatenSheet[i-1]


    #Workbook abspeichern
    wb.save(filename=workbookTitel)
def übungHinzufügen():
    #excel übungsSheet öffnen
    wb = load_workbook(filename = workbookTitel)
    sheetÜbungen = wb['Uebungen']
    sheetDaten = wb['Trainingsstatistiken']

    #Athena fragt die zu hinzufügenden Übungen ab
    athenaSagt("Sehr, gern. Wie ist der Name der Übung?")
    _übung = nutzerSagt()
    athenaSagt("Verstanden. Wie viele Sätze %s möchtest du machen?" % _übung)
    _sets = nutzerSagt()
    athenaSagt("Ok. Wie viele Wiederholungen willst du pro Satz machen?")
    _reps = nutzerSagt()
    athenaSagt("Alles klar. Ich werde nun die Übung in deine Excel-Datei hinzufügen. Sollte etwas nicht stimmen schaue bitte dort nach.")


    liste = [_übung, _sets, _reps]
    sheetÜbungen.append(liste)
    #auf das nächste freie column die attribute eintragen
    maxSpalte = sheetDaten.max_column
    sheetDaten.cell(row =1, column=maxSpalte + 1).value = _übung
    

        
    #excel speichern
    wb.save(filename= workbookTitel)
def löscheWorkbook():
    os.remove(workbookTitel)
def setStarten(_übung, _sets, _reps):
    #öffne workbook
    #wb = load_workbook(filename= workbookTitel)
    #öffne sheet 'sheetÜbungen' und 'Trainingsdaten'
    #sheetÜbungen = wb['Uebungen']
    #sheetDaten = wb['Trainingsstatistiken']
    #Athena sagt wie viele Reps du im Set1 machen sollst und wartet auf die tatsächliche Anzahl an Reps die du ihr sagst
    for setx in range(1, _sets):
        athenaSagt("%s. %s Satz. %s Wiederholungen" % (_übung, str(setx) + '.', _reps))
        athenaSagt("Nachdem du mit der Übung fertig bist, sage bitte wie viele Wiederholungen du geschafft hast. Wecke mich dazu wieder mit: Ok, Athena")
        time.sleep(1)
        nutzerStimme = ''
        
        time.sleep(1)
        while 1:
            nutzerStimme = nutzerSagt()
            

        

    #Du sagst es ihr und die neuen Reps werden in den 'Trainingsdaten' eingetragen an die Stelle der Trainingssession(Datum)
    #Athena bestätigt und wiederholt was sie eintragen wird
    #Excel speichern

def fortschrittEintragen(_reps):
    pass



def starteTraining():
    athenaSagt('Beginnen wir mit dem Training')
    #Workbook und Sheets laden
    wb = load_workbook(filename=workbookTitel)
    sheetÜbungen = wb['Uebungen']
    #sheetDaten = wb['Trainingsstatistiken']
    #Die Gesamtanzahl an Übungen ablesen und als Variable speichern
    gesamtanzahlÜbungen = sheetÜbungen.max_row
    #An die erste Übung gehen, die Anzahl der Sets und Reps abspeichern und setStarten mit den Vars starten
    for stelle in range(2,gesamtanzahlÜbungen):
        übung = sheetÜbungen.cell(row=stelle, column=1).value
        sets = sheetÜbungen.cell(row=stelle, column=2).value
        reps = sheetÜbungen.cell(row=stelle, column=3).value
        setStarten(übung, sets, reps)


    
erstnutzungPrüfen()


time.sleep(0.25)
while 0.25:
    nutzerStimme = nutzerSagt()
    athenaEntscheidet(nutzerStimme)


